"""
Excel CRUD — Créer, lire, modifier et supprimer des fichiers XLSX.

Catégorie: file
Mode: sync

Actions:
    create  → Génère un XLSX depuis des données structurées (feuilles, colonnes, lignes)
    read    → Parse un XLSX et retourne les données structurées
    update  → Modifie un XLSX (ajouter/modifier des cellules, feuilles)
    delete  → Supprime un fichier XLSX du stockage
"""

from __future__ import annotations

import io
from typing import Any

from app.framework.base import BaseTool
from app.framework.schemas import (
    ToolErrorCode,
    ToolExample,
    ToolExecutionMode,
    ToolMetadata,
    ToolParameter,
    ToolResult,
)


class ExcelCrud(BaseTool):
    """CRUD complet pour les fichiers XLSX via MinIO."""

    @property
    def metadata(self) -> ToolMetadata:
        return ToolMetadata(
            slug="excel-crud",
            name="Excel CRUD",
            description="Créer, lire, modifier et supprimer des fichiers Excel (XLSX)",
            version="1.0.0",
            category="file",
            execution_mode=ToolExecutionMode.SYNC,
            timeout_seconds=30,
            input_schema=[
                ToolParameter(name="action", type="string", required=True,
                              description="Action: create, read, update, delete"),
                ToolParameter(name="storage_key", type="string",
                              description="Chemin MinIO du fichier"),
                ToolParameter(name="data", type="object",
                              description="Données: {sheets: [{name, headers, rows}]}"),
                ToolParameter(name="options", type="object",
                              description="Options: {sheet_name, header_row}"),
            ],
            output_schema=[
                ToolParameter(name="storage_key", type="string"),
                ToolParameter(name="sheets", type="array",
                              description="Liste des feuilles avec données"),
                ToolParameter(name="sheet_names", type="array"),
                ToolParameter(name="sheet_count", type="integer"),
            ],
            examples=[
                ToolExample(
                    description="Créer un Excel",
                    input={
                        "action": "create",
                        "storage_key": "reports/ventes.xlsx",
                        "data": {
                            "sheets": [
                                {
                                    "name": "Janvier",
                                    "headers": ["Produit", "Quantité", "Prix"],
                                    "rows": [
                                        ["Widget A", 100, 29.99],
                                        ["Widget B", 50, 49.99],
                                    ],
                                },
                            ],
                        },
                    },
                    output={"storage_key": "reports/ventes.xlsx", "sheet_count": 1},
                ),
                ToolExample(
                    description="Lire un Excel",
                    input={"action": "read", "storage_key": "reports/ventes.xlsx"},
                    output={
                        "sheet_names": ["Janvier"],
                        "sheets": [{"name": "Janvier", "headers": ["Produit"], "rows": [["Exemple"]]}],
                    },
                ),
            ],
            tags=["file", "excel", "xlsx", "crud", "office", "tabular"],
        )

    async def execute(self, params: dict[str, Any], context) -> ToolResult:
        action = params.get("action", "")
        if action == "create":
            return await self._create(params, context)
        elif action == "read":
            return await self._read(params, context)
        elif action == "update":
            return await self._update(params, context)
        elif action == "delete":
            return await self._delete(params, context)
        else:
            return self.error(
                f"Action inconnue: '{action}'. Actions: create, read, update, delete",
                ToolErrorCode.INVALID_PARAMS,
            )

    async def _create(self, params: dict[str, Any], context) -> ToolResult:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        storage_key = params.get("storage_key", "")
        data = params.get("data", {})

        if not storage_key:
            return self.error("storage_key requis pour create", ToolErrorCode.INVALID_PARAMS)

        wb = Workbook()
        # Supprimer la feuille par défaut
        wb.remove(wb.active)

        sheets = data.get("sheets", [])
        if not sheets:
            sheets = [{"name": "Sheet1", "headers": [], "rows": []}]

        for sheet_data in sheets:
            ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])

            # Headers en gras
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=str(header))
                    cell.font = Font(bold=True)

            # Données
            start_row = 2 if headers else 1
            for row_idx, row_data in enumerate(rows, start_row):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Auto-ajuster la largeur des colonnes
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted = min(max_length + 2, 50)
                ws.column_dimensions[col[0].column_letter].width = adjusted

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        await context.storage.put(
            storage_key, buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        return self.success({
            "storage_key": storage_key,
            "sheet_names": [s.get("name", "Sheet") for s in sheets],
            "sheet_count": len(sheets),
        })

    async def _read(self, params: dict[str, Any], context) -> ToolResult:
        from openpyxl import load_workbook

        storage_key = params.get("storage_key", "")
        options = params.get("options", {})

        if not storage_key:
            return self.error("storage_key requis pour read", ToolErrorCode.INVALID_PARAMS)

        file_data = await context.storage.get(storage_key)
        if file_data is None:
            return self.error(f"Fichier introuvable: {storage_key}", ToolErrorCode.FILE_NOT_FOUND)

        try:
            wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        except Exception as e:
            return self.error(f"Fichier XLSX invalide: {e}", ToolErrorCode.PROCESSING_ERROR)

        target_sheet = options.get("sheet_name")
        header_row = options.get("header_row", 1)

        sheets_data = []
        for ws in wb.worksheets:
            if target_sheet and ws.title != target_sheet:
                continue

            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([self._serialize_cell(v) for v in row])

            headers = all_rows[header_row - 1] if len(all_rows) >= header_row else []
            data_rows = all_rows[header_row:] if len(all_rows) > header_row else []

            sheets_data.append({
                "name": ws.title,
                "headers": headers,
                "rows": data_rows,
                "row_count": len(data_rows),
                "column_count": len(headers),
            })

        wb.close()

        return self.success({
            "storage_key": storage_key,
            "sheets": sheets_data,
            "sheet_names": [s["name"] for s in sheets_data],
            "sheet_count": len(sheets_data),
        })

    async def _update(self, params: dict[str, Any], context) -> ToolResult:
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        storage_key = params.get("storage_key", "")
        data = params.get("data", {})

        if not storage_key:
            return self.error("storage_key requis pour update", ToolErrorCode.INVALID_PARAMS)

        file_data = await context.storage.get(storage_key)
        if file_data is None:
            return self.error(f"Fichier introuvable: {storage_key}", ToolErrorCode.FILE_NOT_FOUND)

        try:
            wb = load_workbook(io.BytesIO(file_data))
        except Exception as e:
            return self.error(f"Fichier XLSX invalide: {e}", ToolErrorCode.PROCESSING_ERROR)

        # Modifier des cellules spécifiques
        cell_updates = data.get("cells", {})
        for sheet_name, cells in cell_updates.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell_ref, value in cells.items():
                    ws[cell_ref] = value

        # Ajouter des lignes
        add_rows = data.get("add_rows", {})
        for sheet_name, rows in add_rows.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row_data in rows:
                    ws.append(row_data)

        # Ajouter une nouvelle feuille
        add_sheets = data.get("add_sheets", [])
        for sheet_data in add_sheets:
            ws = wb.create_sheet(title=sheet_data.get("name", "New Sheet"))
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=str(header))
                    cell.font = Font(bold=True)
            start_row = 2 if headers else 1
            for row_idx, row_data in enumerate(rows, start_row):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        await context.storage.put(
            storage_key, buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        return self.success({
            "storage_key": storage_key,
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        })

    async def _delete(self, params: dict[str, Any], context) -> ToolResult:
        storage_key = params.get("storage_key", "")
        if not storage_key:
            return self.error("storage_key requis pour delete", ToolErrorCode.INVALID_PARAMS)

        deleted = await context.storage.delete(storage_key)
        if not deleted:
            return self.error(f"Fichier introuvable: {storage_key}", ToolErrorCode.FILE_NOT_FOUND)

        return self.success({"storage_key": storage_key, "deleted": True})

    @staticmethod
    def _serialize_cell(value: Any) -> Any:
        """Convertit les valeurs de cellules en types JSON-safe."""
        if value is None:
            return None
        if isinstance(value, (int, float, bool, str)):
            return value
        return str(value)
